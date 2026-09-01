"""
Bus Management API

Handles CRUD operations for school buses.
"""

from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from sqlalchemy.orm import Session

from backend.database import get_db
from backend.models import Bus, Driver, FleetNotification, Route, Student
from backend.schemas import BusCreate, BusUpdate, BusResponse
from backend.security import require_management
from backend.routes.models_tracking import (
    BusGPSState,
    GPSDeviceMapping,
    GPSIngestToken,
    LiveLocation,
    LiveTrip,
    TripStopEvent,
    ProviderGPSPosition,
)

# ======================================================
# BUILD BUS RESPONSE
# ======================================================

def build_bus_response(bus: Bus, db: Session) -> BusResponse:

    driver_name = None

    if bus.driver_id:

        driver = (
            db.query(Driver)
            .filter(Driver.id == bus.driver_id)
            .first()
        )

        if driver and driver.user:

            driver_name = driver.user.full_name

    return BusResponse(

        id=bus.id,

        bus_number=bus.bus_number,
        registration_number=bus.registration_number,
        capacity=bus.capacity,

        manufacturer=bus.manufacturer,
        model=bus.model,
        year=bus.year,

        fuel_type=bus.fuel_type,
        status=bus.status,

        driver_id=bus.driver_id,
        driver_name=driver_name,

        route=bus.route,
        device_id=bus.device_id,

        created_at=bus.created_at,
        updated_at=bus.updated_at,
    )




router = APIRouter(
    prefix="/api/buses",
    tags=["Bus Management"]
)


BUS_EXPORT_HEADERS = [
    "Index",
    "Bus Number",
    "Registration Number",
    "Capacity",
    "Manufacturer",
    "Model",
    "Year",
    "Fuel Type",
    "GPS Device ID",
    "Status",
]


@router.get("/", response_model=list[BusResponse])
def get_buses(
    db: Session = Depends(get_db),
    _current_user = Depends(require_management),
):

    buses = db.query(Bus).order_by(Bus.bus_number).all()

    return [

        build_bus_response(bus, db)

        for bus in buses

    ]


@router.get("/export")
def export_buses(
    db: Session = Depends(get_db),
    _current_user = Depends(require_management),
):
    """Export every field maintained by the Add Bus form to Excel."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Buses"
    sheet.append(BUS_EXPORT_HEADERS)

    buses = db.query(Bus).order_by(Bus.bus_number.asc()).all()
    for index, bus in enumerate(buses, start=1):
        sheet.append([
            index,
            bus.bus_number,
            bus.registration_number,
            bus.capacity,
            bus.manufacturer,
            bus.model,
            bus.year,
            bus.fuel_type,
            bus.device_id,
            bus.status,
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Buses.xlsx"'},
    )


@router.post("/import")
async def import_buses(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _current_user = Depends(require_management),
):
    """Import an exported Bus workbook without duplicating fleet records."""

    try:
        workbook = load_workbook(BytesIO(await file.read()), data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail="Please upload a valid Excel (.xlsx) file.") from error

    sheet = workbook.active
    headers = {
        str(cell.value).strip().casefold(): index
        for index, cell in enumerate(sheet[1])
        if cell.value is not None and str(cell.value).strip()
    }
    required_headers = {header.casefold() for header in BUS_EXPORT_HEADERS[1:]}
    missing_headers = sorted(required_headers - headers.keys())
    if missing_headers:
        raise HTTPException(
            status_code=400,
            detail="The Excel sheet is missing required Bus export columns.",
        )

    def value_for(row: tuple, header: str):
        index = headers[header.casefold()]
        return row[index] if index < len(row) else None

    def clean_text(value) -> str:
        return "" if value is None else str(value).strip()

    buses = db.query(Bus).all()
    buses_by_number = {bus.bus_number.strip().casefold(): bus for bus in buses}
    buses_by_registration = {bus.registration_number.strip().casefold(): bus for bus in buses}
    imported_buses: list[dict] = []
    skipped_buses: list[dict] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and clean_text(value) for value in row):
            continue

        payload = {
            "bus_number": clean_text(value_for(row, "Bus Number")),
            "registration_number": clean_text(value_for(row, "Registration Number")),
            "capacity": value_for(row, "Capacity"),
            "manufacturer": clean_text(value_for(row, "Manufacturer")),
            "model": clean_text(value_for(row, "Model")),
            "year": value_for(row, "Year"),
            "fuel_type": clean_text(value_for(row, "Fuel Type")),
            "device_id": clean_text(value_for(row, "GPS Device ID")) or None,
            "status": clean_text(value_for(row, "Status")) or "Active",
        }
        try:
            validated = BusCreate.model_validate(payload)
        except ValidationError as error:
            skipped_buses.append({
                "row": row_number,
                "bus_number": payload["bus_number"] or None,
                "registration_number": payload["registration_number"] or None,
                "reason": error.errors()[0]["msg"],
            })
            continue

        number_match = buses_by_number.get(validated.bus_number.casefold())
        registration_match = buses_by_registration.get(validated.registration_number.casefold())
        if number_match is not None and registration_match is not None and number_match.id != registration_match.id:
            skipped_buses.append({
                "row": row_number,
                "bus_number": validated.bus_number,
                "registration_number": validated.registration_number,
                "reason": "Bus number and registration number match different existing buses",
            })
            continue
        existing_bus = number_match or registration_match
        if existing_bus is not None:
            skipped_buses.append({
                "row": row_number,
                "bus_number": existing_bus.bus_number,
                "registration_number": existing_bus.registration_number,
                "reason": "Bus already exists",
            })
            continue

        bus = Bus(**validated.model_dump())
        db.add(bus)
        db.flush()
        buses_by_number[bus.bus_number.casefold()] = bus
        buses_by_registration[bus.registration_number.casefold()] = bus
        imported_buses.append({
            "row": row_number,
            "bus_number": bus.bus_number,
            "registration_number": bus.registration_number,
        })

    db.commit()
    return {
        "success": True,
        "imported": len(imported_buses),
        "skipped": len(skipped_buses),
        "summary": {
            "imported_buses": imported_buses,
            "skipped_buses": skipped_buses,
        },
        "message": "Bus import completed.",
    }

@router.post(
    "/",
    response_model=BusResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_bus(
    bus: BusCreate,
    db: Session = Depends(get_db),
    _current_user = Depends(require_management),
):
    """Create a new bus."""

    existing_bus = (
        db.query(Bus)
        .filter(Bus.bus_number == bus.bus_number)
        .first()
    )

    if existing_bus:
        raise HTTPException(
            status_code=400,
            detail=(
                f'Bus number "{bus.bus_number}" is already registered '
                f'as {existing_bus.bus_number} (registration '
                f'"{existing_bus.registration_number}"). Choose a different '
                "bus number or edit the existing bus."
            ),
        )

    existing_registration = (
        db.query(Bus)
        .filter(Bus.registration_number == bus.registration_number)
        .first()
    )

    if existing_registration:
        raise HTTPException(
            status_code=400,
            detail=(
                f'Registration number "{bus.registration_number}" is already '
                f"used by bus {existing_registration.bus_number}. Choose a "
                "different registration number or edit that bus."
            ),
        )

    new_bus = Bus(
        bus_number=bus.bus_number,
        registration_number=bus.registration_number,
        capacity=bus.capacity,
        manufacturer=bus.manufacturer,
        model=bus.model,
        year=bus.year,
        fuel_type=bus.fuel_type,
        status=bus.status,
        device_id=bus.device_id,
    )

    db.add(new_bus)
    db.flush()

    
    db.commit()
    db.refresh(new_bus)

    return build_bus_response(new_bus, db)

@router.get("/{bus_id}", response_model=BusResponse)
def get_bus(
    bus_id: int,
    db: Session = Depends(get_db),
    _current_user = Depends(require_management),
):
    """Return a single bus."""

    bus = db.query(Bus).filter(Bus.id == bus_id).first()

    if not bus:
        raise HTTPException(
            status_code=404,
            detail="Bus not found.",
        )
    return build_bus_response(bus, db)

    

@router.put("/{bus_id}", response_model=BusResponse)
def update_bus(
    bus_id: int,
    bus: BusUpdate,
    db: Session = Depends(get_db),
    _current_user = Depends(require_management),
):
    """Update a bus."""

    existing = (
        db.query(Bus)
        .filter(Bus.id == bus_id)
        .first()
    )

    if not existing:
        raise HTTPException(
            status_code=404,
            detail="Bus not found.",
        )

    duplicate_bus = (
        db.query(Bus)
        .filter(
            Bus.bus_number == bus.bus_number,
            Bus.id != bus_id,
        )
        .first()
    )

    if duplicate_bus:
        raise HTTPException(
            status_code=400,
            detail=(
                f'Bus number "{bus.bus_number}" is already registered '
                f'as {duplicate_bus.bus_number} (registration '
                f'"{duplicate_bus.registration_number}"). Choose a different '
                "bus number or edit the existing bus."
            ),
        )

    duplicate_registration = (
        db.query(Bus)
        .filter(
            Bus.registration_number == bus.registration_number,
            Bus.id != bus_id,
        )
        .first()
    )

    if duplicate_registration:
        raise HTTPException(
            status_code=400,
            detail=(
                f'Registration number "{bus.registration_number}" is already '
                f"used by bus {duplicate_registration.bus_number}. Choose a "
                "different registration number or edit that bus."
            ),
        )

    existing.bus_number = bus.bus_number
    existing.registration_number = bus.registration_number
    existing.capacity = bus.capacity
    existing.manufacturer = bus.manufacturer
    existing.model = bus.model
    existing.year = bus.year
    existing.fuel_type = bus.fuel_type
    existing.status = bus.status
    existing.device_id = bus.device_id

    db.commit()
    db.refresh(existing)


    return build_bus_response(existing, db)

@router.delete("/{bus_id}")
def delete_bus(
    bus_id: int,
    db: Session = Depends(get_db),
    _current_user = Depends(require_management),
):
    """Delete a bus."""

    bus = (
        db.query(Bus)
        .filter(Bus.id == bus_id)
        .first()
    )

    if not bus:
        raise HTTPException(
            status_code=404,
            detail="Bus not found."
        )

    # Clear every record that points at this bus before deleting it.
    db.query(Route).filter(Route.bus_id == bus.id).update(
        {Route.bus_id: None, Route.driver_id: None}, synchronize_session=False
    )
    db.query(Driver).filter(Driver.bus_id == bus.id).update(
        {Driver.bus_id: None}, synchronize_session=False
    )
    db.query(Student).filter(Student.bus_id == bus.id).update(
        {Student.bus_id: None}, synchronize_session=False
    )
    # Live trips retain a required bus reference, and each location retains a
    # required trip reference. Remove the dependent locations first so a bus
    # with historical or active tracking data can be deleted safely.
    trip_ids = db.query(LiveTrip.id).filter(LiveTrip.bus_id == bus.id)
    # Detach feedback before deleting its associated trip as well. This is
    # explicit so it also works with older database schemas lacking ON DELETE
    # SET NULL constraints.
    db.query(FleetNotification).filter(FleetNotification.trip_id.in_(trip_ids)).update(
        {FleetNotification.trip_id: None}, synchronize_session=False
    )
    db.query(TripStopEvent).filter(TripStopEvent.trip_id.in_(trip_ids)).delete(
        synchronize_session=False
    )
    db.query(LiveLocation).filter(LiveLocation.trip_id.in_(trip_ids)).delete(
        synchronize_session=False
    )
    db.query(LiveTrip).filter(LiveTrip.bus_id == bus.id).delete(
        synchronize_session=False
    )
    # Preserve operational feedback, but detach it from the deleted vehicle.
    db.query(FleetNotification).filter(FleetNotification.bus_id == bus.id).update(
        {FleetNotification.bus_id: None}, synchronize_session=False
    )
    bus.driver_id = None
    bus.route = None

    # Provider identities and their history belong to the deleted bus. Clear
    # them explicitly because the established schema does not use DB-level
    # cascade rules for bus references.
    db.query(BusGPSState).filter(BusGPSState.bus_id == bus.id).delete(synchronize_session=False)
    db.query(ProviderGPSPosition).filter(ProviderGPSPosition.bus_id == bus.id).delete(synchronize_session=False)
    db.query(GPSDeviceMapping).filter(GPSDeviceMapping.bus_id == bus.id).delete(synchronize_session=False)
    db.query(GPSIngestToken).filter(GPSIngestToken.bus_id == bus.id).delete(synchronize_session=False)

    db.delete(bus)

    db.commit()

    return {
        "message": "Bus deleted successfully."
    }
