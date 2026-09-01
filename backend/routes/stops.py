"""
==========================================================
BUSTRACK
MASTER STOPS API
==========================================================

This API manages the master list of transport stops.

Routes DO NOT own stops.
Routes reference stops through the RouteStop table.
"""

# ==========================================================
# IMPORTS
# ==========================================================
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
import io
import math
from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    UploadFile,
    File,
)
from sqlalchemy.orm import Session, selectinload
from openpyxl import load_workbook
from io import BytesIO
from backend.database import get_db
from backend.models import (
    Stop,
    Route,
    RouteStop,
)
from backend.security import require_management
# ==========================================================
# ROUTER
# ==========================================================

router = APIRouter(
    prefix="/api/stops",
    tags=["Stops"],
)
# ==========================================================
# GET ALL STOPS
# ==========================================================

@router.get("")
def get_stops(
    db: Session = Depends(get_db),
    _current_user = Depends(require_management),
):
    """
    Returns the master list of all transport stops.

    Stops are independent entities and can be reused
    across multiple routes.
    """

    stops = (

        db.query(Stop)

        .order_by(
            Stop.stop_name.asc()
        )

        .all()

    )

    data = []

    for stop in stops:

        data.append({

            "id": stop.id,

            "stop_code": stop.stop_code,

            "stop_name": stop.stop_name,

            "latitude": stop.latitude,

            "longitude": stop.longitude,

            "radius": stop.radius,

            "status": stop.status,

        })

    return data
# ==========================================================
# EXPORT STOPS TO EXCEL
# ==========================================================

@router.get("/export")
def export_stops(

    db: Session = Depends(get_db),
    _current_user = Depends(require_management),

):
    """
    Export all master stops to an Excel file.

    Each row represents one master stop and includes its geofence settings.
    Stops are reusable across routes, so all linked route codes, names, and
    stop sequences are collected into the same row rather than duplicating
    the stop for every route assignment.
    """

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Stops"

    # ======================================================
    # HEADER
    # ======================================================

    sheet.append([
        "Index",
        "Stop Code",
        "Stop Name",
        "Latitude",
        "Longitude",
        "Geofence Radius (m)",
        "Status",
        "Route Number(s)",
        "Route Name(s)",
        "Route Stop Sequence(s)",
    ])

    # ======================================================
    # DATA
    # ======================================================

    stops = (

        db.query(Stop)
        .options(
            selectinload(Stop.route_stops).selectinload(RouteStop.route),
        )
        .order_by(Stop.stop_name.asc())
        .all()

    )

    for index, stop in enumerate(stops, start=1):

        route_assignments = sorted(
            (
                route_stop
                for route_stop in stop.route_stops
                if route_stop.route is not None
            ),
            key=lambda route_stop: (
                route_stop.route.route_code,
                route_stop.sequence,
            ),
        )

        sheet.append([
            index,
            stop.stop_code,
            stop.stop_name,
            stop.latitude,
            stop.longitude,
            stop.radius,
            stop.status,
            "; ".join(route_stop.route.route_code for route_stop in route_assignments),
            "; ".join(route_stop.route.route_name for route_stop in route_assignments),
            "; ".join(str(route_stop.sequence) for route_stop in route_assignments),
        ])

    # ======================================================
    # SAVE TO MEMORY
    # ======================================================

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    # ======================================================
    # DOWNLOAD
    # ======================================================

    return StreamingResponse(

        output,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={

            "Content-Disposition":

            'attachment; filename="Stops.xlsx"'

        }

    )   
# ==========================================================
# GET SINGLE STOP
# ==========================================================

@router.get("/{stop_id}")
def get_stop(

    stop_id: int,

    db: Session = Depends(get_db),
    _current_user = Depends(require_management),

):
    """
    Returns a single stop from the master stop database.
    """

    stop = (

        db.query(Stop)

        .filter(
            Stop.id == stop_id
        )

        .first()

    )

    if stop is None:

        raise HTTPException(

            status_code=404,

            detail="Stop not found."

        )

    return {

        "id": stop.id,

        "stop_code": stop.stop_code,

        "stop_name": stop.stop_name,

        "latitude": stop.latitude,

        "longitude": stop.longitude,

        "radius": stop.radius,

        "status": stop.status,

    }
# ==========================================================
# CREATE STOP
# ==========================================================

@router.post("")
def create_stop(

    stop_data: dict,

    db: Session = Depends(get_db),
    _current_user = Depends(require_management),

):
    """
    Creates a new master stop.
    """

    # ======================================================
    # VALIDATE STOP CODE
    # ======================================================

    existing_code = (

        db.query(Stop)

        .filter(
            Stop.stop_code == stop_data["stop_code"]
        )

        .first()

    )

    if existing_code:

        raise HTTPException(

            status_code=400,

            detail="Stop code already exists."

        )

    # ======================================================
    # VALIDATE STOP NAME
    # ======================================================

    existing_name = (

        db.query(Stop)

        .filter(
            Stop.stop_name == stop_data["stop_name"]
        )

        .first()

    )

    if existing_name:

        raise HTTPException(

            status_code=400,

            detail="Stop name already exists."

        )

    # ======================================================
    # CREATE STOP
    # ======================================================

    stop = Stop(

        stop_code = stop_data["stop_code"],

        stop_name = stop_data["stop_name"],

        latitude = stop_data.get("latitude"),

        longitude = stop_data.get("longitude"),

        radius = stop_data.get("radius", 50),

        status = stop_data.get("status", "Active"),

    )

    db.add(stop)

    db.commit()

    db.refresh(stop)

    return {

        "success": True,

        "message": "Stop created successfully.",

        "stop": {

            "id": stop.id,

            "stop_code": stop.stop_code,

            "stop_name": stop.stop_name,

            "latitude": stop.latitude,

            "longitude": stop.longitude,

            "radius": stop.radius,

            "status": stop.status,

        }

    }
# ==========================================================
# UPDATE STOP
# ==========================================================

@router.put("/{stop_id}")
def update_stop(

    stop_id: int,

    stop_data: dict,

    db: Session = Depends(get_db),
    _current_user = Depends(require_management),

):
    """
    Updates an existing master stop.
    """

    stop = (

        db.query(Stop)

        .filter(
            Stop.id == stop_id
        )

        .first()

    )

    if stop is None:

        raise HTTPException(

            status_code=404,

            detail="Stop not found."

        )

    # ======================================================
    # CHECK DUPLICATE STOP CODE
    # ======================================================

    existing_code = (

        db.query(Stop)

        .filter(
            Stop.stop_code == stop_data["stop_code"],
            Stop.id != stop_id
        )

        .first()

    )

    if existing_code:

        raise HTTPException(

            status_code=400,

            detail="Stop code already exists."

        )

    # ======================================================
    # CHECK DUPLICATE STOP NAME
    # ======================================================

    existing_name = (

        db.query(Stop)

        .filter(
            Stop.stop_name == stop_data["stop_name"],
            Stop.id != stop_id
        )

        .first()

    )

    if existing_name:

        raise HTTPException(

            status_code=400,

            detail="Stop name already exists."

        )

    # ======================================================
    # UPDATE VALUES
    # ======================================================

    stop.stop_code = stop_data["stop_code"]

    stop.stop_name = stop_data["stop_name"]

    stop.latitude = stop_data.get("latitude")

    stop.longitude = stop_data.get("longitude")

    stop.radius = stop_data.get("radius", 50)

    stop.status = stop_data.get("status", "Active")

    db.commit()

    db.refresh(stop)

    return {

        "success": True,

        "message": "Stop updated successfully.",

        "stop": {

            "id": stop.id,

            "stop_code": stop.stop_code,

            "stop_name": stop.stop_name,

            "latitude": stop.latitude,

            "longitude": stop.longitude,

            "radius": stop.radius,

            "status": stop.status,

        }

    }
# ==========================================================
# IMPORT STOPS FROM EXCEL
# ==========================================================

@router.post("/import")
async def import_stops(

    file: UploadFile = File(...),

    db: Session = Depends(get_db),
    _current_user = Depends(require_management),

):
    """Import either the legacy two-column sheet or the complete Stops export.

    The complete export is safe to re-import: matching stop codes or names
    are not duplicated. Existing stops can still be connected to the routes
    listed in the upload, preserving the export's route-stop configuration.
    """
    try:
        workbook = load_workbook(BytesIO(await file.read()), data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail="Please upload a valid Excel (.xlsx) file.") from error

    sheet = workbook.active
    headers = {
        str(cell.value).strip().casefold(): index
        for index, cell in enumerate(sheet[1])
        if cell.value is not None and str(cell.value).strip()
    }
    is_complete_export = "stop code" in headers and "stop name" in headers
    if not is_complete_export and "place" not in headers:
        raise HTTPException(
            status_code=400,
            detail="The Excel sheet must contain either Place or Stop Code and Stop Name headers.",
        )

    def get_value(row: tuple, header: str):
        index = headers.get(header.casefold())
        return row[index] if index is not None and index < len(row) else None

    def text_value(value) -> str:
        return "" if value is None else str(value).strip()

    def split_values(value) -> list[str]:
        return [part.strip() for part in text_value(value).split(";") if part.strip()]

    def optional_coordinate(value, label: str) -> float | None:
        if value is None or text_value(value) == "":
            return None
        try:
            coordinate = float(value)
        except (TypeError, ValueError) as error:
            raise ValueError(f"{label} is not a number") from error
        if not math.isfinite(coordinate):
            raise ValueError(f"{label} is not valid")
        if label == "Latitude" and not -90 <= coordinate <= 90:
            raise ValueError("Latitude must be between -90 and 90")
        if label == "Longitude" and not -180 <= coordinate <= 180:
            raise ValueError("Longitude must be between -180 and 180")
        return coordinate

    def optional_radius(value) -> int:
        if value is None or text_value(value) == "":
            return 50
        try:
            radius = int(value)
        except (TypeError, ValueError) as error:
            raise ValueError("Geofence Radius (m) must be a whole number") from error
        if not 10 <= radius <= 500:
            raise ValueError("Geofence Radius (m) must be between 10 and 500")
        return radius

    existing_stops = db.query(Stop).all()
    stops_by_code = {stop.stop_code.strip().casefold(): stop for stop in existing_stops}
    stops_by_name = {stop.stop_name.strip().casefold(): stop for stop in existing_stops}
    routes = db.query(Route).all()
    routes_by_code = {route.route_code.strip().casefold(): route for route in routes}
    routes_by_name = {route.route_name.strip().casefold(): route for route in routes}
    existing_route_stops = {
        (route_stop.route_id, route_stop.stop_id)
        for route_stop in db.query(RouteStop).all()
    }

    highest_generated_stop_number = 0
    for stop in existing_stops:
        code = stop.stop_code.strip().upper()
        if code.startswith("ST") and code[2:].isdigit():
            highest_generated_stop_number = max(highest_generated_stop_number, int(code[2:]))

    imported_stops: list[dict] = []
    skipped_stops: list[dict] = []
    created_routes: list[dict] = []
    linked_route_stops: list[dict] = []
    skipped_route_stops: list[dict] = []
    affected_route_ids: set[int] = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        stop_name = text_value(get_value(row, "Stop Name" if is_complete_export else "Place"))
        if not stop_name:
            continue

        stop_code = text_value(get_value(row, "Stop Code")) if is_complete_export else ""
        try:
            latitude = optional_coordinate(get_value(row, "Latitude"), "Latitude")
            longitude = optional_coordinate(get_value(row, "Longitude"), "Longitude")
            if (latitude is None) != (longitude is None):
                raise ValueError("Latitude and Longitude must both be provided")
            radius = optional_radius(get_value(row, "Geofence Radius (m)"))
        except ValueError as error:
            skipped_stops.append({
                "row": row_number,
                "stop_code": stop_code or None,
                "stop_name": stop_name,
                "reason": str(error),
            })
            continue

        by_code = stops_by_code.get(stop_code.casefold()) if stop_code else None
        by_name = stops_by_name.get(stop_name.casefold())
        if by_code is not None and by_name is not None and by_code.id != by_name.id:
            skipped_stops.append({
                "row": row_number,
                "stop_code": stop_code or None,
                "stop_name": stop_name,
                "reason": "Stop code and stop name match different existing stops",
            })
            continue

        stop = by_code or by_name
        if stop is not None:
            skipped_stops.append({
                "row": row_number,
                "stop_code": stop.stop_code,
                "stop_name": stop.stop_name,
                "reason": "Stop already exists",
            })
        else:
            if not stop_code:
                highest_generated_stop_number += 1
                stop_code = f"ST{highest_generated_stop_number:04d}"
            if len(stop_code) > 20 or len(stop_name) > 150:
                skipped_stops.append({
                    "row": row_number,
                    "stop_code": stop_code,
                    "stop_name": stop_name,
                    "reason": "Stop code or stop name is too long",
                })
                continue
            stop = Stop(
                stop_code=stop_code,
                stop_name=stop_name,
                latitude=latitude,
                longitude=longitude,
                radius=radius,
                status=text_value(get_value(row, "Status")) or "Active",
            )
            db.add(stop)
            db.flush()
            stops_by_code[stop.stop_code.casefold()] = stop
            stops_by_name[stop.stop_name.casefold()] = stop
            imported_stops.append({"row": row_number, "stop_code": stop.stop_code, "stop_name": stop.stop_name})

        if not is_complete_export:
            continue

        route_codes = split_values(get_value(row, "Route Number(s)"))
        route_names = split_values(get_value(row, "Route Name(s)"))
        route_sequences = split_values(get_value(row, "Route Stop Sequence(s)"))
        if not route_codes and not route_names and not route_sequences:
            continue
        if not (len(route_codes) == len(route_names) == len(route_sequences)):
            skipped_route_stops.append({
                "row": row_number,
                "stop_name": stop.stop_name,
                "reason": "Route number, name, and sequence counts must match",
            })
            continue

        for route_code, route_name, sequence_value in zip(route_codes, route_names, route_sequences):
            try:
                sequence = int(sequence_value)
                if sequence < 1:
                    raise ValueError
            except (TypeError, ValueError):
                skipped_route_stops.append({
                    "row": row_number,
                    "stop_name": stop.stop_name,
                    "route_code": route_code,
                    "reason": "Route stop sequence must be a positive whole number",
                })
                continue

            route_by_code = routes_by_code.get(route_code.casefold())
            route_by_name = routes_by_name.get(route_name.casefold())
            if route_by_code is not None and route_by_name is not None and route_by_code.id != route_by_name.id:
                skipped_route_stops.append({
                    "row": row_number,
                    "stop_name": stop.stop_name,
                    "route_code": route_code,
                    "reason": "Route number and route name match different existing routes",
                })
                continue
            route = route_by_code or route_by_name
            if route is None:
                if not route_code or not route_name or len(route_code) > 20 or len(route_name) > 100:
                    skipped_route_stops.append({
                        "row": row_number,
                        "stop_name": stop.stop_name,
                        "route_code": route_code or None,
                        "reason": "A valid route number and route name are required to create a route",
                    })
                    continue
                route = Route(route_code=route_code, route_name=route_name, status="Active")
                db.add(route)
                db.flush()
                routes_by_code[route.route_code.casefold()] = route
                routes_by_name[route.route_name.casefold()] = route
                created_routes.append({"route_code": route.route_code, "route_name": route.route_name})

            route_stop_key = (route.id, stop.id)
            if route_stop_key in existing_route_stops:
                skipped_route_stops.append({
                    "row": row_number,
                    "stop_name": stop.stop_name,
                    "route_code": route.route_code,
                    "reason": "Stop is already linked to this route",
                })
                continue

            db.add(RouteStop(route_id=route.id, stop_id=stop.id, sequence=sequence))
            existing_route_stops.add(route_stop_key)
            affected_route_ids.add(route.id)
            linked_route_stops.append({
                "row": row_number,
                "stop_name": stop.stop_name,
                "route_code": route.route_code,
                "sequence": sequence,
            })

    db.flush()
    for route_id in affected_route_ids:
        route = db.get(Route, route_id)
        route.total_stops = db.query(RouteStop).filter(RouteStop.route_id == route_id).count()
    db.commit()

    return {
        "success": True,
        "imported": len(imported_stops),
        "skipped": len(skipped_stops),
        "routes_created": len(created_routes),
        "route_stops_linked": len(linked_route_stops),
        "route_stops_skipped": len(skipped_route_stops),
        "summary": {
            "imported_stops": imported_stops,
            "skipped_stops": skipped_stops,
            "created_routes": created_routes,
            "linked_route_stops": linked_route_stops,
            "skipped_route_stops": skipped_route_stops,
        },
        "message": "Stops import completed.",
    }
# ==========================================================
# DELETE STOP
# ==========================================================

# ==========================================================
# DELETE STOP
# ==========================================================

@router.delete("/{stop_id}")
def delete_stop(

    stop_id: int,

    db: Session = Depends(get_db),
    _current_user = Depends(require_management),

):
    """
    Deletes a stop from the master stop database.
    """

    stop = (

        db.query(Stop)

        .filter(
            Stop.id == stop_id
        )

        .first()

    )

    if stop is None:

        raise HTTPException(

            status_code=404,

            detail="Stop not found."

        )

    # ======================================================
    # CHECK WHETHER STOP IS USED IN A ROUTE
    # ======================================================

    existing_route = (

        db.query(RouteStop)

        .filter(
            RouteStop.stop_id == stop.id
        )

        .first()

    )

    if existing_route:

        raise HTTPException(

            status_code=400,

            detail="This stop is used in one or more routes."

        )

    # ======================================================
    # DELETE STOP
    # ======================================================

    db.delete(stop)

    db.commit()

    return {

        "success": True,

        "message": "Stop deleted successfully."

    }
