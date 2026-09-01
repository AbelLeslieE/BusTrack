"""Regression coverage for the Bus Excel export/import round trip."""

from __future__ import annotations

import asyncio
from io import BytesIO
from pathlib import Path
import tempfile
import unittest
from uuid import uuid4

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import backend.models  # noqa: F401
import backend.routes.models_tracking  # noqa: F401
from backend.database import Base
from backend.models import Bus
from backend.routes.buses import BUS_EXPORT_HEADERS, export_buses, import_buses


class BusImportExportTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.test_database = Path(tempfile.gettempdir()) / f"bus_tracker_bus_export_{uuid4().hex}.db"
        cls.engine = create_engine(f"sqlite:///{cls.test_database.as_posix()}")
        cls.session_factory = sessionmaker(bind=cls.engine, autoflush=False)
        Base.metadata.create_all(bind=cls.engine)

    @classmethod
    def tearDownClass(cls) -> None:
        Base.metadata.drop_all(bind=cls.engine)
        cls.engine.dispose()
        cls.test_database.unlink(missing_ok=True)

    def setUp(self) -> None:
        with self.session_factory() as database_session:
            database_session.query(Bus).delete()
            database_session.commit()

    def test_export_includes_every_add_bus_field(self) -> None:
        with self.session_factory() as database_session:
            bus = Bus(
                bus_number="BUS-010",
                registration_number="KL-10-AB-1234",
                capacity=48,
                manufacturer="Ashok Leyland",
                model="Lynx Smart",
                year=2025,
                fuel_type="Diesel",
                device_id="GPS-1010",
                status="Maintenance",
            )
            database_session.add(bus)
            database_session.commit()

            response = export_buses(database_session, object())
            content = asyncio.run(self._read_response(response))

        sheet = load_workbook(BytesIO(content), data_only=True)["Buses"]
        self.assertEqual([cell.value for cell in sheet[1]], BUS_EXPORT_HEADERS)
        self.assertEqual(
            [cell.value for cell in sheet[2]],
            [1, "BUS-010", "KL-10-AB-1234", 48, "Ashok Leyland", "Lynx Smart", 2025,
             "Diesel", "GPS-1010", "Maintenance"],
        )

    def test_import_creates_full_bus_records_and_skips_duplicate_uploads(self) -> None:
        with self.session_factory() as database_session:
            database_session.add(Bus(
                bus_number="BUS-001",
                registration_number="KL-01-AA-0001",
                capacity=40,
                manufacturer="Existing",
                model="Coach",
                year=2024,
                fuel_type="Diesel",
                status="Active",
            ))
            database_session.commit()

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(BUS_EXPORT_HEADERS)
            sheet.append([1, "BUS-002", "KL-01-AA-0002", 52, "Tata", "Starbus", 2026, "CNG", "GPS-002", "Active"])
            sheet.append([2, "BUS-001", "KL-01-AA-0001", 40, "Existing", "Coach", 2024, "Diesel", None, "Active"])
            contents = BytesIO()
            workbook.save(contents)

            result = asyncio.run(import_buses(
                UploadFile(filename="buses.xlsx", file=BytesIO(contents.getvalue())),
                database_session,
                object(),
            ))
            self.assertEqual((result["imported"], result["skipped"]), (1, 1))
            imported = database_session.query(Bus).filter(Bus.bus_number == "BUS-002").one()
            self.assertEqual(
                (imported.registration_number, imported.capacity, imported.manufacturer, imported.model,
                 imported.year, imported.fuel_type, imported.device_id, imported.status),
                ("KL-01-AA-0002", 52, "Tata", "Starbus", 2026, "CNG", "GPS-002", "Active"),
            )

            duplicate_result = asyncio.run(import_buses(
                UploadFile(filename="buses.xlsx", file=BytesIO(contents.getvalue())),
                database_session,
                object(),
            ))
            self.assertEqual((duplicate_result["imported"], duplicate_result["skipped"]), (0, 2))
            self.assertEqual(database_session.query(Bus).count(), 2)

    @staticmethod
    async def _read_response(response) -> bytes:
        return b"".join([chunk async for chunk in response.body_iterator])


if __name__ == "__main__":
    unittest.main()
