"""Regression coverage for the complete master-stop Excel export."""

from __future__ import annotations

import asyncio
from io import BytesIO
from pathlib import Path
import tempfile
import unittest
from uuid import uuid4

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import backend.models  # noqa: F401
import backend.routes.models_tracking  # noqa: F401
from backend.database import Base
from backend.models import Route, RouteStop, Stop
from backend.routes.stops import export_stops, import_stops


class StopExportTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.test_database = Path(tempfile.gettempdir()) / f"bus_tracker_stop_export_{uuid4().hex}.db"
        cls.engine = create_engine(f"sqlite:///{cls.test_database.as_posix()}")
        cls.session_factory = sessionmaker(bind=cls.engine, autoflush=False)
        Base.metadata.create_all(bind=cls.engine)

    @classmethod
    def tearDownClass(cls) -> None:
        Base.metadata.drop_all(bind=cls.engine)
        cls.engine.dispose()
        cls.test_database.unlink(missing_ok=True)

    def setUp(self) -> None:
        with self.session_factory() as database_session:
            database_session.query(RouteStop).delete()
            database_session.query(Stop).delete()
            database_session.query(Route).delete()
            database_session.commit()

    def test_export_includes_geofence_details_and_all_route_assignments(self) -> None:
        with self.session_factory() as database_session:
            stop = Stop(
                stop_code="ST010",
                stop_name="Market Junction",
                latitude=10.123456,
                longitude=76.654321,
                radius=125,
                status="Active",
            )
            route_one = Route(route_code="R-01", route_name="Morning Route", status="Active")
            route_two = Route(route_code="R-02", route_name="Evening Route", status="Active")
            database_session.add_all([stop, route_one, route_two])
            database_session.flush()
            database_session.add_all([
                RouteStop(route_id=route_two.id, stop_id=stop.id, sequence=4),
                RouteStop(route_id=route_one.id, stop_id=stop.id, sequence=2),
            ])
            database_session.commit()

            response = export_stops(database_session, object())
            content = asyncio.run(self._read_response(response))

        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook["Stops"]
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            [
                "Index", "Stop Code", "Stop Name", "Latitude", "Longitude",
                "Geofence Radius (m)", "Status", "Route Number(s)",
                "Route Name(s)", "Route Stop Sequence(s)",
            ],
        )
        self.assertEqual(
            [cell.value for cell in sheet[2]],
            [
                1, "ST010", "Market Junction", 10.123456, 76.654321, 125,
                "Active", "R-01; R-02", "Morning Route; Evening Route", "2; 4",
            ],
        )

    @staticmethod
    async def _read_response(response) -> bytes:
        return b"".join([chunk async for chunk in response.body_iterator])

    def test_complete_export_format_imports_geofences_routes_and_skips_duplicates(self) -> None:
        with self.session_factory() as database_session:
            existing_stop = Stop(stop_code="ST001", stop_name="Existing Stop", radius=50, status="Active")
            database_session.add(existing_stop)
            database_session.commit()

            workbook = Workbook()
            sheet = workbook.active
            sheet.append([
                "Index", "Stop Code", "Stop Name", "Latitude", "Longitude",
                "Geofence Radius (m)", "Status", "Route Number(s)",
                "Route Name(s)", "Route Stop Sequence(s)",
            ])
            sheet.append([1, "ST100", "New Stop", 10.123456, 76.654321, 125, "Active", "R-100", "New Route", "1"])
            sheet.append([2, "ST001", "Existing Stop", 10.2, 76.7, 80, "Active", "R-100", "New Route", "2"])
            contents = BytesIO()
            workbook.save(contents)

            result = asyncio.run(import_stops(
                UploadFile(filename="stops.xlsx", file=BytesIO(contents.getvalue())),
                database_session,
                object(),
            ))
            self.assertEqual(result["imported"], 1)
            self.assertEqual(result["skipped"], 1)
            self.assertEqual(result["routes_created"], 1)
            self.assertEqual(result["route_stops_linked"], 2)

            new_stop = database_session.query(Stop).filter(Stop.stop_code == "ST100").one()
            self.assertEqual((new_stop.latitude, new_stop.longitude, new_stop.radius), (10.123456, 76.654321, 125))
            route = database_session.query(Route).filter(Route.route_code == "R-100").one()
            self.assertEqual(route.total_stops, 2)
            self.assertEqual(
                [item.sequence for item in database_session.query(RouteStop).filter(RouteStop.route_id == route.id).order_by(RouteStop.sequence)],
                [1, 2],
            )

            duplicate_result = asyncio.run(import_stops(
                UploadFile(filename="stops.xlsx", file=BytesIO(contents.getvalue())),
                database_session,
                object(),
            ))
            self.assertEqual(duplicate_result["imported"], 0)
            self.assertEqual(duplicate_result["skipped"], 2)
            self.assertEqual(duplicate_result["route_stops_linked"], 0)
            self.assertEqual(duplicate_result["route_stops_skipped"], 2)
            self.assertEqual(database_session.query(Stop).count(), 2)
            self.assertEqual(database_session.query(RouteStop).count(), 2)

    def test_legacy_place_export_remains_supported(self) -> None:
        with self.session_factory() as database_session:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Index", "Place"])
            sheet.append([1, "Legacy Stop"])
            contents = BytesIO()
            workbook.save(contents)

            result = asyncio.run(import_stops(
                UploadFile(filename="legacy-stops.xlsx", file=BytesIO(contents.getvalue())),
                database_session,
                object(),
            ))

            self.assertEqual(result["imported"], 1)
            stop = database_session.query(Stop).one()
            self.assertEqual(stop.stop_name, "Legacy Stop")
            self.assertEqual(stop.stop_code, "ST0001")
            self.assertEqual(stop.radius, 50)


if __name__ == "__main__":
    unittest.main()
